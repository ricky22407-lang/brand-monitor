"""
routes/export.py
匯出功能：/api/export/excel
"""

import os
from datetime import datetime
from flask import Blueprint, jsonify, send_file
from core.state import state

bp = Blueprint("export", __name__)


@bp.route("/api/export/excel")
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "輿情報表"

        headers    = ["類型","來源","版區","所屬文章","內容","作者","時間","情緒","信心度","AI摘要","連結"]
        col_widths = [8,     10,    10,    40,        40,    16,    18,    8,    8,       30,      60]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1C2030")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[chr(64 + col)].width = w

        for row, a in enumerate(state["articles"], 2):
            is_comment = a.get("type") == "comment"
            kind       = "留言" if is_comment else "文章"
            sent_label = {"positive": "正面", "negative": "負面", "neutral": "中性"}.get(a.get("sentiment", ""), "")
            score_str  = f"{round(a.get('sentiment_score', 0) * 100)}%" if a.get("sentiment_score") else ""

            vals = [
                kind,
                a.get("source", ""),
                a.get("board",  ""),
                a.get("parent_title", "") if is_comment else "",
                a.get("content", "")[:300] if is_comment else a.get("title", ""),
                a.get("author",    ""),
                a.get("timestamp", "")[:16],
                sent_label,
                score_str,
                a.get("sentiment_summary", ""),
                a.get("url", ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # 情緒底色
            sent = a.get("sentiment", "")
            ws.cell(row=row, column=8).fill = PatternFill("solid",
                fgColor="C6EFCE" if sent == "positive" else ("FFC7CE" if sent == "negative" else "FFEB9C"))
            # 類型底色
            ws.cell(row=row, column=1).fill = PatternFill("solid",
                fgColor="D9EAD3" if not is_comment else "CFE2F3")

        ws.freeze_panes = "A2"

        os.makedirs("exports", exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = f"exports/輿情報表_{ts}.xlsx"
        wb.save(path)
        return send_file(path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
